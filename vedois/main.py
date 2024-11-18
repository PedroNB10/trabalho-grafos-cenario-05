import os
import csv
import pandas as pd
import networkx as nx
import matplotlib.pyplot as plt
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import random

# adicionei a verificação em "check_conflicts" para impedir que o curso de CCO tenha disciplinas alocadas no periodo noturno
# adicionei a verificação em "generate_schedules" para impedir que um professor ultrapasse as 6 horas de aula por dia (no máximo 6 aulas)

# Criar diretórios caso não existam
os.makedirs('dataset', exist_ok=True)
os.makedirs('controle', exist_ok=True)
os.makedirs('professores (bruto)', exist_ok=True)
os.makedirs('professores (importar)', exist_ok=True)
os.makedirs('alunos (importar)', exist_ok=True)
os.makedirs('grafos', exist_ok=True) 

# Ler arquivo da pasta dataset
df = pd.read_csv(os.path.join('dataset', 'semestre1.csv'))

def create_assignment_graphs():
    assignment_graph_name = nx.Graph()
    assignment_graph_code = nx.Graph()
    
    # Criar um dicionário para rastrear ocorrências múltiplas
    disc_occurrences = defaultdict(int)
    code_occurrences = defaultdict(int)
    
    for _, row in df.iterrows():
        # Incrementar contadores
        disc_occurrences[row['Nome da Disciplina']] += 1
        code_occurrences[row['Código da Disciplina']] += 1
        
        # Adicionar sufixo para disciplinas duplicadas
        disciplina_nome = f"{row['Nome da Disciplina']}_{disc_occurrences[row['Nome da Disciplina']]}"
        disciplina_codigo = f"{row['Código da Disciplina']}_{code_occurrences[row['Código da Disciplina']]}"
        
        for i in range(1, 19):
            prof_col = f'Prof {i}'
            if row[prof_col] == 1:
                assignment_graph_name.add_edge(f'Prof {i}', disciplina_nome)
                assignment_graph_code.add_edge(f'Prof {i}', disciplina_codigo)
    
    return assignment_graph_name, assignment_graph_code

def create_constraint_graphs():
    constraint_graph_name = nx.Graph()
    constraint_graph_code = nx.Graph()
    prof_to_disc = defaultdict(list)
    
    # Rastrear ocorrências múltiplas
    disc_occurrences = defaultdict(int)
    code_occurrences = defaultdict(int)
    
    for _, row in df.iterrows():
        # Incrementar contadores
        disc_occurrences[row['Nome da Disciplina']] += 1
        code_occurrences[row['Código da Disciplina']] += 1
        
        disc_nome = f"{row['Nome da Disciplina']}_{disc_occurrences[row['Nome da Disciplina']]}"
        disc_codigo = f"{row['Código da Disciplina']}_{code_occurrences[row['Código da Disciplina']]}"
        periodo = row['Período']
        curso = row['Curso']
        
        constraint_graph_name.add_node(disc_nome)
        constraint_graph_code.add_node(disc_codigo)
        
        for i in range(1, 19):
            if row[f'Prof {i}'] == 1:
                prof_to_disc[f'Prof {i}'].append((disc_nome, disc_codigo))
    
        # Adicionar arestas para disciplinas do mesmo período/curso
        for _, other_row in df.iterrows():
            other_disc_occurrences = disc_occurrences[other_row['Nome da Disciplina']]
            other_code_occurrences = code_occurrences[other_row['Código da Disciplina']]
            
            if (other_row['Período'] == periodo and 
                other_row['Curso'] == curso and 
                other_row['Nome da Disciplina'] != row['Nome da Disciplina']):
                other_disc_nome = f"{other_row['Nome da Disciplina']}_{other_disc_occurrences}"
                other_disc_codigo = f"{other_row['Código da Disciplina']}_{other_code_occurrences}"
                constraint_graph_name.add_edge(disc_nome, other_disc_nome)
                constraint_graph_code.add_edge(disc_codigo, other_disc_codigo)
    
    # Adicionar arestas entre disciplinas do mesmo professor
    for prof_discs in prof_to_disc.values():
        for i in range(len(prof_discs)):
            for j in range(i + 1, len(prof_discs)):
                constraint_graph_name.add_edge(prof_discs[i][0], prof_discs[j][0])
                constraint_graph_code.add_edge(prof_discs[i][1], prof_discs[j][1])
    
    return constraint_graph_name, constraint_graph_code

def check_conflicts(schedule, new_allocation):
    """
    Verifica se há conflitos de horário para a nova alocação
    """
    dia, turno, horario = new_allocation['Dia'], new_allocation['Turno'], new_allocation['Horário']
    curso, periodo = new_allocation['Curso'], new_allocation['Período']
    professor = new_allocation['Professor']
    
    # Verificar restrição de horários por curso
    if curso == 'CCO' and turno == 'N':  # CCO não pode ter aulas à noite
        return True
    if curso == 'SIN' and turno != 'N':  # SIN só pode ter aulas à noite
        return True
    
    # Verificar limite de 6 horas por dia para professor
    prof_aulas_dia = schedule[
        (schedule['Dia'] == dia) & 
        (schedule['Professor'] == professor)
    ]
    if len(prof_aulas_dia) >= 6:
        return True
    
    # Verificar conflitos no mesmo horário
    conflitos = schedule[
        (schedule['Dia'] == dia) & 
        (schedule['Turno'] == turno) & 
        (schedule['Horário'] == horario)
    ]
    
    # Se houver alguma aula no mesmo horário
    for _, aula in conflitos.iterrows():
        # Verificar conflito de professor
        if aula['Professor'] == professor:
            return True
        # Verificar conflito de turma (mesmo período/curso)
        if aula['Curso'] == curso and aula['Período'] == periodo:
            return True
    
    return False

def get_saturation_degree(vertex, graph, colored_vertices):
    """Calcula o grau de saturação de um vértice."""
    # Conjunto de cores diferentes usadas pelos vizinhos
    neighbor_colors = set()
    for neighbor in graph.neighbors(vertex):
        if neighbor in colored_vertices:
            neighbor_colors.add(colored_vertices[neighbor])
    return len(neighbor_colors)

def get_max_saturation_vertex(graph, colored_vertices, uncolored_vertices):
    """Retorna o vértice não colorido com maior grau de saturação."""
    max_saturation = -1
    max_degree = -1
    selected_vertex = None
    
    for vertex in uncolored_vertices:
        saturation = get_saturation_degree(vertex, graph, colored_vertices)
        degree = graph.degree(vertex)
        
        # Se encontrarmos uma saturação maior ou igual com maior grau
        if (saturation > max_saturation) or \
           (saturation == max_saturation and degree > max_degree):
            max_saturation = saturation
            max_degree = degree
            selected_vertex = vertex
            
    return selected_vertex

def dsatur_coloring(graph):
    """Implementa o algoritmo DSATUR para coloração do grafo."""
    colored_vertices = {}  # Dicionário para armazenar as cores dos vértices
    uncolored_vertices = set(graph.nodes())  # Conjunto de vértices não coloridos
    
    # Enquanto houver vértices não coloridos
    while uncolored_vertices:
        # Seleciona o vértice com maior grau de saturação
        vertex = get_max_saturation_vertex(graph, colored_vertices, uncolored_vertices)
        
        # Encontra a menor cor disponível
        used_colors = set(colored_vertices[neighbor] 
                         for neighbor in graph.neighbors(vertex) 
                         if neighbor in colored_vertices)
        
        # Encontra a primeira cor não utilizada pelos vizinhos
        color = 0
        while color in used_colors:
            color += 1
            
        # Atribui a cor ao vértice
        colored_vertices[vertex] = color
        uncolored_vertices.remove(vertex)
    
    return colored_vertices

def generate_schedules():
    MAX_ATTEMPTS = 1000  # Número máximo de tentativas para alocar uma disciplina
    schedules = {}
    dias = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta']
    turnos = ['M', 'T', 'N']
    horarios = list(range(1, 6))
    
    # Criar um DataFrame global para rastrear todas as alocações
    global_schedule = pd.DataFrame(columns=['Dia', 'Turno', 'Horário', 'Curso',
                                          'Período', 'Código', 'Disciplina', 'CH', 'Professor'])
    
    # Criar arquivo de log para disciplinas não alocadas
    log_path = "disciplinas_nao_alocadas.log"
    with open(log_path, "w") as log_file:
        log_file.write("Disciplinas não alocadas completamente:\n")
    
    # Ordenar disciplinas por restrições e carga horária
    disciplinas = []
    for _, row in df.iterrows():
        profs = sum([1 for i in range(1, 19) if row[f'Prof {i}'] == 1])
        disciplinas.append({
            'row': row,
            'restricoes': profs + (2 if row['Curso'] in ['CCO', 'SIN'] else 0) + (row['CH'] / 2)
        })
    
    # Ordenar por número de restrições e CH (mais restritivas primeiro)
    disciplinas.sort(key=lambda x: x['restricoes'], reverse=True)
    
    def find_sequential_slot(prof, row, turno, remaining_ch):
        """
        Tenta encontrar slots sequenciais para a disciplina no mesmo dia e turno
        """
        for dia in dias:
            prof_aulas_dia = len(global_schedule[
                (global_schedule['Dia'] == dia) & 
                (global_schedule['Professor'] == prof)
            ])
            
            if prof_aulas_dia >= 6:
                continue
                
            # Verificar sequência de horários disponíveis
            for horario_inicial in horarios:
                if horario_inicial + remaining_ch > 6:
                    continue
                    
                pode_alocar = True
                new_allocations = []
                
                for h in range(remaining_ch):
                    new_allocation = {
                        'Dia': dia,
                        'Turno': turno,
                        'Horário': horario_inicial + h,
                        'Curso': row['Curso'],
                        'Período': row['Período'],
                        'Código': row['Código da Disciplina'],
                        'Disciplina': row['Nome da Disciplina'],
                        'CH': row['CH'],
                        'Professor': prof
                    }
                    
                    if check_conflicts(global_schedule, new_allocation):
                        pode_alocar = False
                        break
                        
                    new_allocations.append(new_allocation)
                
                if pode_alocar:
                    return new_allocations
                    
        return None

    # Alocar disciplinas
    for disc in disciplinas:
        row = disc['row']
        ch = row['CH']
        curso = row['Curso']
        
        # Identificar professor(es)
        profs = [f'Prof {i}' for i in range(1, 19) if row[f'Prof {i}'] == 1]
        
        # Definir turnos baseado no curso
        if curso == 'CCO':
            turnos_pref = ['M', 'T']  # CCO só pode ter aulas de manhã e tarde
        elif curso == 'SIN':
            turnos_pref = ['N']  # SIN só pode ter aulas à noite
        else:
            turnos_pref = turnos  # Outros cursos podem ter aulas em qualquer turno
        
        # Alocar aulas
        remaining_ch = ch
        total_attempts = 0
        
        while remaining_ch > 0 and total_attempts < MAX_ATTEMPTS:
            allocated = False
            
            # Tentar primeiro alocar em sequência
            for prof in profs:
                for turno in turnos_pref:
                    if allocated:
                        break
                        
                    sequential_slots = find_sequential_slot(prof, row, turno, min(2, remaining_ch))
                    if sequential_slots:
                        for allocation in sequential_slots:
                            global_schedule = pd.concat([global_schedule, 
                                                       pd.DataFrame([allocation])], 
                                                      ignore_index=True)
                        allocated = True
                        remaining_ch -= len(sequential_slots)
                        break
            
            # Se não foi possível alocar em sequência, tenta alocar individualmente
            if not allocated:
                random.shuffle(dias)
                random.shuffle(horarios)
                
                for prof in profs:
                    if allocated:
                        break
                        
                    for dia in dias:
                        if allocated:
                            break
                            
                        for turno in turnos_pref:
                            if allocated:
                                break
                                
                            # Verificar quantidade de aulas do professor no dia
                            prof_aulas_dia = len(global_schedule[
                                (global_schedule['Dia'] == dia) & 
                                (global_schedule['Professor'] == prof)
                            ])
                            
                            if prof_aulas_dia >= 6:
                                continue
                                
                            for horario in horarios:
                                new_allocation = {
                                    'Dia': dia,
                                    'Turno': turno,
                                    'Horário': horario,
                                    'Curso': curso,
                                    'Período': row['Período'],
                                    'Código': row['Código da Disciplina'],
                                    'Disciplina': row['Nome da Disciplina'],
                                    'CH': row['CH'],
                                    'Professor': prof
                                }
                                
                                if not check_conflicts(global_schedule, new_allocation):
                                    global_schedule = pd.concat([global_schedule, 
                                                               pd.DataFrame([new_allocation])], 
                                                              ignore_index=True)
                                    allocated = True
                                    remaining_ch -= 1
                                    break
            
            if not allocated:
                total_attempts += 1
            
            if total_attempts >= MAX_ATTEMPTS:
                print(f"AVISO: Não foi possível alocar completamente a disciplina {row['Nome da Disciplina']} após {MAX_ATTEMPTS} tentativas")
                break
    
    # Separar schedules por professor
    for i in range(1, 19):
        prof = f'Prof {i}'
        schedules[prof] = global_schedule[global_schedule['Professor'] == prof].copy()
    
    # Verificar disciplinas não alocadas completamente
    with open(log_path, "a") as log_file:
        for disc in disciplinas:
            row = disc['row']
            aulas_alocadas = len(global_schedule[global_schedule['Código'] == row['Código da Disciplina']])
            if aulas_alocadas < row['CH']:
                msg = (f"Disciplina {row['Nome da Disciplina']} ({row['Código da Disciplina']}) - "
                       f"CH necessária: {row['CH']}, CH alocada: {aulas_alocadas}\n")
                print(msg.strip())
                log_file.write(msg)
    
    return schedules, global_schedule

def plot_graph(graph, title, colors=None):
    plt.figure(figsize=(15, 10))
    pos = nx.spring_layout(graph, k=1, iterations=50)
    
    if colors:
        node_colors = [colors.get(node, 'white') for node in graph.nodes()]
        nx.draw(graph, pos, with_labels=True, node_color=node_colors,
               node_size=1500, font_size=6, font_weight='bold')
    else:
        nx.draw(graph, pos, with_labels=True,
               node_size=1500, font_size=6, font_weight='bold')
    
    plt.title(title)
    
    # Salvar o gráfico na pasta 'grafos'
    plt.savefig(os.path.join('grafos', f'{title}.png'))  # Salva o gráfico com o título como nome do arquivo
    
    # Exibir o gráfico
    plt.show()

def display_professor_schedule(schedules):
    """
    Exibe o horário de cada professor em formato organizado
    e salva os dados em um arquivo CSV.
    """
    dias = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta']
    turnos = ['M', 'T', 'N']
    horarios = range(1, 6)

    # Lista para armazenar dados que serão salvos no CSV
    csv_data = []
    
    for i in range(1, 19):
        prof = f'Prof {i}'
        print(f"\nHorário do {prof}:")
        print("-" * 80)
        
        prof_schedule = schedules[prof]
        
        for dia in dias:
            print(f"\n{dia}:")
            for turno in turnos:
                print(f"\nTurno {turno}:")
                for horario in horarios:
                    aula = prof_schedule[
                        (prof_schedule['Dia'] == dia) & 
                        (prof_schedule['Turno'] == turno) & 
                        (prof_schedule['Horário'] == horario)
                    ]
                    
                    if len(aula) > 0:
                        row = aula.iloc[0]
                        disciplina = f"{row['Disciplina']} ({row['Curso']} - {row['Período']}º período)"
                        print(f"Horário {horario}: {disciplina}")
                    else:
                        disciplina = "livre"
                        print(f"Horário {horario}: livre")

                    # Adiciona os dados na lista para o CSV
                    csv_data.append({
                        'Professor': prof,
                        'Dia': dia,
                        'Turno': turno,
                        'Horário': horario,
                        'Disciplina': disciplina
                    })
        print("=" * 80)
    
    # Salva os dados no arquivo CSV dentro da pasta 'controle'
    with open('controle/planilha.csv', mode='w', newline='', encoding='utf-8') as file:
        writer = csv.DictWriter(file, fieldnames=['Professor', 'Dia', 'Turno', 'Horário', 'Disciplina'])
        writer.writeheader()
        writer.writerows(csv_data)

def save_schedules(schedules, global_schedule):
    # Salvar arquivo global
    global_schedule.to_csv(os.path.join('controle', 'horario_global.csv'), index=False)
    
    # Salvar grades individuais
    for i in range(1, 19):
        prof = f'Prof {i}'
        prof_schedule = schedules[prof]
        prof_schedule.to_csv(os.path.join('professores (bruto)', f'{prof}_schedule.csv'), index=False)
    
    print("Horários salvos com sucesso.")

# Executar as funções
assignment_graph_name, assignment_graph_code = create_assignment_graphs()
constraint_graph_name, constraint_graph_code = create_constraint_graphs()

# Aplicar coloração DSATUR
colors_name = dsatur_coloring(constraint_graph_name)
colors_code = dsatur_coloring(constraint_graph_code)

# Mostrar quantidade de cores utilizadas
num_colors_name = len(set(colors_name.values()))
num_colors_code = len(set(colors_code.values()))
print(f"Número de cores utilizadas (grafo de nomes): {num_colors_name}")
print(f"Número de cores utilizadas (grafo de códigos): {num_colors_code}")

# Gerar visualizações
plot_graph(assignment_graph_name, "Grafo de Atribuição (Nomes das Disciplinas)")
plot_graph(assignment_graph_code, "Grafo de Atribuição (Códigos das Disciplinas)")
plot_graph(constraint_graph_name, "Grafo de Restrições (Nomes das Disciplinas)", colors_name)
plot_graph(constraint_graph_code, "Grafo de Restrições (Códigos das Disciplinas)", colors_code)

schedules, global_schedule = generate_schedules()
display_professor_schedule(schedules)
save_schedules(schedules, global_schedule)

# GERA ARQUIVOS IMPORTÁVEIS PARA SPREADSHEET OU EXCEL (PROFESSORES)
# Dicionário para mapear turno e horário
# Serve para professores e alunos
horarios = {
    "M1": "07:00 - 07:55", "M2": "07:55 - 08:50", "M3": "08:50 - 09:45",
    "M4": "10:10 - 11:00", "M5": "11:00 - 11:55", "T1": "13:30 - 14:25",
    "T2": "14:25 - 15:20", "T3": "15:45 - 16:40", "T4": "16:40 - 17:35",
    "T5": "17:35 - 18:30", "N1": "19:00 - 19:50", "N2": "19:50 - 20:40",
    "N3": "21:00 - 21:50", "N4": "21:50 - 22:40", "N5": "22:40 - 23:30"
}

# Carregar o arquivo CSV
arquivo_csv = "controle/planilha.csv"
df = pd.read_csv(arquivo_csv)

# Substituir os identificadores pelos horários no DataFrame
df["Turno_Horário"] = (df["Turno"].astype(str) + df["Horário"].astype(str)).map(horarios)

# Agrupar os dados por professor
professores = df.groupby("Professor")

# Cores para as células
red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
green_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")

print("Planilhas de professores geradas com sucesso!")

# GERA ARQUIVOS IMPORTÁVEIS PARA SPREADSHEET OU EXCEL (ALUNOS)
# Função para concatenar os horários
def get_horario(dia, turno, horario):
    # Construir a chave para o dicionário de horários
    chave = f"{turno}{horario}"
    # Se o horário existir no dicionário, retornamos o valor correspondente
    if chave in horarios:
        return f"{dia} {turno} {horarios[chave]}"
    else:
        return "Desconhecido"

print("Planilhas de alunos geradas com sucesso!")

def format_cell(cell, bg_color=None, font_color="000000", bold=False, merge_cells=None, alignment=None):
    """Helper function to format cells"""
    if bg_color:
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.font = Font(color=font_color, bold=bold)
    if alignment:
        cell.alignment = alignment
    if merge_cells:
        worksheet = cell.parent
        worksheet.merge_cells(merge_cells)

def create_schedule_workbook(data, is_professor=False):
    wb = Workbook()
    ws = wb.active

    # Define styles
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers setup
    if is_professor:
        # Professor header
        prof_name = data['Professor'].iloc[0]
        cell = ws.cell(row=1, column=1, value=prof_name)
        format_cell(cell, 
                   bg_color="000000", 
                   font_color="FFFFFF", 
                   bold=True, 
                   merge_cells=f'A1:F1',
                   alignment=center_alignment)
    else:
        # Course header
        course = f"{data['Curso'].iloc[0]} / {data['Período'].iloc[0]}º Período"
        cell = ws.cell(row=1, column=1, value=course)
        format_cell(cell, 
                   bg_color="000000", 
                   font_color="FFFFFF", 
                   bold=True, 
                   merge_cells=f'A1:F1',
                   alignment=center_alignment)

    # Day headers
    days = ['Horário', 'Segunda-feira', 'Terça-feira', 'Quarta-feira', 'Quinta-feira', 'Sexta-feira']
    for col, day in enumerate(days, 1):
        cell = ws.cell(row=2, column=col, value=day)
        format_cell(cell, bold=True, alignment=center_alignment)
        cell.border = border

    # Time slots
    time_slots = {
        'M': ['07:00 - 07:55', '07:55 - 08:50', '08:50 - 09:45', '10:10 - 11:00', '11:00 - 11:55'],
        'T': ['13:30 - 14:25', '14:25 - 15:20', '15:45 - 16:40', '16:40 - 17:35', '17:35 - 18:30'],
        'N': ['19:00 - 19:50', '19:50 - 20:40', '21:00 - 21:50', '21:50 - 22:40', '22:40 - 23:30']
    }

    current_row = 3
    for turn in ['M', 'T', 'N']:
        for idx, time in enumerate(time_slots[turn]):
            # Time column
            cell = ws.cell(row=current_row, column=1, value=time)
            format_cell(cell, alignment=center_alignment)
            cell.border = border

            # Initialize empty cells for each day
            for day_col in range(2, 7):
                cell = ws.cell(row=current_row, column=day_col, value="")
                cell.border = border

            current_row += 1

    # Fill in the schedule
    colors = {
        'XMAC': 'CC99FF',  # Purple
        'XDES': 'CCCCCC',  # Gray
        'MAT': 'FFFF99',   # Yellow
        'CRSC': '99CCFF',  # Light Blue
        'IEPG': 'FF99CC',  # Pink
        'SAHC': 'FF99FF',  # Light Purple
        'CAHC': 'FF0000',  # Red
    }

    # Process each class
    for _, row in data.iterrows():
        day_map = {
            'Segunda': 2,
            'Terça': 3,
            'Quarta': 4,
            'Quinta': 5,
            'Sexta': 6
        }
        
        time_map = {
            ('M', 1): 3,
            ('M', 2): 4,
            ('M', 3): 5,
            ('M', 4): 6,
            ('M', 5): 7,
            ('T', 1): 8,
            ('T', 2): 9,
            ('T', 3): 10,
            ('T', 4): 11,
            ('T', 5): 12,
            ('N', 1): 13,
            ('N', 2): 14,
            ('N', 3): 15,
            ('N', 4): 16,
            ('N', 5): 17
        }

        day_col = day_map.get(row['Dia'])
        time_row = time_map.get((row['Turno'], row['Horário']))

        if day_col and time_row:
            # Determine cell color based on course code
            color_key = next((k for k in colors.keys() if row['Código'].startswith(k)), 'FFFFFF')
            bg_color = colors.get(color_key, 'FFFFFF')

            # Format cell content
            if is_professor:
                content = f"{row['Curso']} {row['Período']}º\n{row['Código']}\n{row['Disciplina']}"
            else:
                content = f"{row['Código']}\n{row['Disciplina']}"

            cell = ws.cell(row=time_row, column=day_col, value=content)
            format_cell(cell, bg_color=bg_color, alignment=center_alignment)
            cell.border = border

    # Adjust column widths
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 20

    return wb

def generate_new_schedules(global_schedule_path):
    # Read the global schedule
    df = pd.read_csv(global_schedule_path)
    
    # Generate professor schedules
    for prof in df['Professor'].unique():
        prof_data = df[df['Professor'] == prof]
        wb = create_schedule_workbook(prof_data, is_professor=True)
        wb.save(f'professores (importar)/{prof}_horarios.xlsx')

    # Generate course schedules
    for (curso, periodo), group_data in df.groupby(['Curso', 'Período']):
        wb = create_schedule_workbook(group_data, is_professor=False)
        wb.save(f'alunos (importar)/{curso}_{periodo}.xlsx')

    print("Novas planilhas geradas com sucesso!")

# Call the function
generate_new_schedules('controle/horario_global.csv')