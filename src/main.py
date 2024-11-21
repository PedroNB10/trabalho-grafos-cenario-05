import csv
import os
import random
import logging
import networkx as nx
import matplotlib.pyplot as plt
from collections import defaultdict
from typing import List, Dict, Optional
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

from classes.Disciplina import Disciplina

logging.basicConfig(filename='erros_agendamento.log', level=logging.ERROR,
                    format='%(asctime)s - %(levelname)s - %(message)s')

def verificar_restricoes_professor(horarios: List[Dict], nos: List[Disciplina]) -> bool:
    horarios_professores = defaultdict(lambda: defaultdict(list))
    
    for dia in range(len(horarios)):
        for turno, cor in horarios[dia].items():
            if cor is not None:
                # Encontrar disciplinas com esta cor
                disciplinas_no_turno = [no for no in nos if no.cor == cor]
                
                # Verificar horário de cada professor
                for disciplina in disciplinas_no_turno:
                    for prof in disciplina.professores:
                        horarios_professores[prof][f"{dia}_{turno}"].append(disciplina)
                        
                        # Verificar se professor tem mais de uma disciplina no mesmo horário
                        if len(horarios_professores[prof][f"{dia}_{turno}"]) > 1:
                            logging.error(f"Professor {prof} tem múltiplas disciplinas em {dia}_{turno}")
                            return False
    
    return True

def verificar_restricoes_turma(horarios: List[Dict], nos: List[Disciplina]) -> bool:
    horarios_turmas = defaultdict(lambda: defaultdict(list))
    
    for dia in range(len(horarios)):
        for turno, cor in horarios[dia].items():
            if cor is not None:
                # Encontrar disciplinas com esta cor
                disciplinas_no_turno = [no for no in nos if no.cor == cor]
                
                # Verificar horário de cada turma
                for disciplina in disciplinas_no_turno:
                    chave_turma = (disciplina.turma, disciplina.curso, disciplina.periodo)
                    horarios_turmas[f"{dia}_{turno}"][chave_turma].append(disciplina)
                    
                    # Verificar se turma tem múltiplas disciplinas no mesmo horário
                    if len(horarios_turmas[f"{dia}_{turno}"][chave_turma]) > 1:
                        logging.error(f"Turma {chave_turma} tem múltiplas disciplinas em {dia}_{turno}")
                        return False
    
    return True

def verificar_restricoes_curso(horarios: List[Dict], nos: List[Disciplina]) -> bool:
    for disciplina in nos:
        if disciplina.curso == 'SIN' and disciplina.horario and not disciplina.horario.startswith('N'):
            logging.error(f"Disciplina de SIN {disciplina.nome} não agendada à noite")
            return False
        
        if disciplina.curso == 'CCO' and disciplina.horario and disciplina.horario.startswith('N'):
            logging.error(f"Disciplina de CCO {disciplina.nome} agendada à noite")
            return False
    
    return True

def salvar_grafo_como_imagem(grafo: nx.Graph):
    """
    Saves the graph visualization with enhanced layout and styling.
    
    Args:
        grafo (nx.Graph): NetworkX graph to visualize
    """
    plt.figure(figsize=(15, 12))
    
    # Use different layout algorithms based on graph size
    if len(grafo) > 50:
        pos = nx.spring_layout(grafo, k=0.8, iterations=50)
    else:
        pos = nx.kamada_kawai_layout(grafo)
    
    # Draw nodes by type with different colors
    node_colors = [grafo.nodes[node].get('color', 'gray') for node in grafo.nodes()]
    node_sizes = [
        3000 if grafo.nodes[node]['type'] == 'disciplina'
        else 1500  # for professors
        for node in grafo.nodes()
    ]
    
    # Draw nodes
    nx.draw_networkx_nodes(
        grafo, pos,
        node_color=node_colors,
        node_size=node_sizes,
        alpha=0.7
    )
    
    # Draw edges with varying weights
    edge_weights = [grafo[u][v].get('weight', 1) for u, v in grafo.edges()]
    nx.draw_networkx_edges(
        grafo, pos,
        width=edge_weights,
        alpha=0.4,
        edge_color='gray'
    )
    
    # Add labels with custom formatting
    labels = {}
    for node in grafo.nodes():
        if grafo.nodes[node]['type'] == 'disciplina':
            labels[node] = f"{node}\n({grafo.nodes[node]['codigo']})"
        else:
            labels[node] = node
            
    nx.draw_networkx_labels(
        grafo, pos,
        labels,
        font_size=8,
        font_weight='bold'
    )
    
    plt.title("Grafo de Relações entre Disciplinas e Professores", 
              fontsize=16, pad=20)
    plt.axis('off')
    plt.tight_layout()
    
    # Save with high resolution
    plt.savefig("grafo_disciplina_curso.png", dpi=300, bbox_inches='tight')
    plt.close()

def salvar_grafo_restricoes_como_imagem(grafo: nx.Graph, caminho: str = "grafo_restricoes.png"):
    """
    Salva o grafo de restrições como uma imagem com cores e layout otimizado.
    
    Args:
        grafo: NetworkX Graph contendo o grafo de restrições
        caminho: Caminho onde a imagem será salva
    """
    plt.figure(figsize=(15, 12))
    
    # Usa um layout que minimiza o cruzamento de arestas
    pos = nx.spring_layout(grafo, k=2, iterations=50)
    
    # Extrai as cores dos nós para visualização
    cores = [grafo.nodes[node]['cor'] for node in grafo.nodes()]
    
    # Desenha os nós
    nx.draw_networkx_nodes(
        grafo,
        pos,
        node_color=cores,
        node_size=2000,
        cmap=plt.cm.tab20,
        alpha=0.7
    )
    
    # Desenha as arestas
    nx.draw_networkx_edges(
        grafo,
        pos,
        edge_color='gray',
        alpha=0.2
    )
    
    # Adiciona os labels dos nós
    labels = {node: f"{node}\n({grafo.nodes[node]['curso']}-{grafo.nodes[node]['periodo']})"
             for node in grafo.nodes()}
    nx.draw_networkx_labels(
        grafo,
        pos,
        labels,
        font_size=8,
        font_weight='bold'
    )
    
    plt.title("Grafo de Restrições de Agendamento", fontsize=16, pad=20)
    plt.axis('off')
    plt.tight_layout()
    
    # Salva a imagem
    plt.savefig(caminho, dpi=300, bbox_inches='tight')
    plt.close()

def criar_grafo_disciplina_curso(nos: List[Disciplina]) -> nx.Graph:
    """
    Creates a comprehensive graph representing relationships between disciplines, courses, and professors.
    
    Args:
        nos (List[Disciplina]): List of Disciplina objects containing course information
        
    Returns:
        nx.Graph: A NetworkX graph with disciplines, courses, periods, and professors as nodes
        
    Features:
        - Separates nodes by type (discipline, course, professor, period)
        - Adds relevant attributes to nodes (credits, course code, etc.)
        - Creates meaningful connections between related entities
        - Uses different node colors for different entity types
        - Includes edge weights based on relationships
    """
    G = nx.Graph()
    
    # Track unique entities
    cursos = set()
    periodos = set()
    
    # Add nodes with attributes
    for disciplina in nos:
        # Add discipline node with full attributes
        G.add_node(
            disciplina.nome,
            type='disciplina',
            curso=disciplina.curso,
            periodo=disciplina.periodo,
            codigo=disciplina.codigo,
            ch=disciplina.ch,
            color='lightblue',
            professors=disciplina.professores
        )
        
        # Add professor nodes and connect to discipline
        for prof in disciplina.professores:
            prof_node = f"Prof: {prof}"
            if not G.has_node(prof_node):
                G.add_node(
                    prof_node,
                    type='professor',
                    color='lightpink'
                )

            # Connect professor to discipline
            G.add_edge(
                disciplina.nome,
                prof_node,
                weight=1,
                relationship='teaches'
            )
    
    return G

def criar_grafo_coloracao_restricoes(nos: List[Disciplina], arestas: Dict) -> nx.Graph:
    """
    Cria um grafo NetworkX que representa as restrições de coloração entre disciplinas.
    
    Args:
        nos: Lista de objetos Disciplina
        arestas: Dicionário de adjacência representando as restrições
    
    Returns:
        NetworkX Graph com as disciplinas e suas restrições
    """
    G = nx.Graph()
    
    # Adiciona nós com atributos
    for disciplina in nos:
        G.add_node(
            disciplina.nome,
            cor=disciplina.cor,
            curso=disciplina.curso,
            periodo=disciplina.periodo,
            professores=','.join(map(str, disciplina.professores))
        )
    
    # Adiciona arestas baseadas nas restrições
    for i, disciplina_i in enumerate(nos):
        for j in arestas[i]:
            disciplina_j = nos[j]
            G.add_edge(
                disciplina_i.nome,
                disciplina_j.nome,
                tipo='restricao'
            )
    
    return G

def exportar_horarios(horarios: List[Dict], nos: List[Disciplina]):
    # Criar pastas se não existirem
    os.makedirs("professor/csv", exist_ok=True)
    os.makedirs("controle", exist_ok=True)
    os.makedirs("aluno/csv", exist_ok=True)

    # Exportar horários por curso/turma
    horarios_curso = defaultdict(list)

    for disciplina in nos:
        chave_curso = f"{disciplina.curso}_{disciplina.periodo}"
        for dia in range(5):  # Semana tem 5 dias
            for turno in horarios[dia].keys():
                if horarios[dia][turno] == disciplina.cor:
                    horarios_curso[chave_curso].append({
                        'Curso': disciplina.curso,
                        'Período': disciplina.periodo,
                        'Código da Disciplina': disciplina.codigo,
                        'Nome da Disciplina': disciplina.nome,
                        'Dia da Semana': dia + 1,  # Semana tem 5 dias uteis
                        'Professor': ', '.join(map(str, disciplina.professores)),
                        'Horário': turno
                    })

    # Escrever horários das turmas
    for chave_curso, horario in horarios_curso.items():
        with open(os.path.join("aluno/csv", f"horario_{chave_curso}.csv"), 'w', newline='', encoding='utf-8') as f:
            escritor = csv.DictWriter(f, fieldnames=['Curso', 'Período', 'Código da Disciplina', 
                                                     'Nome da Disciplina', 'Dia da Semana', 
                                                     'Professor', 'Horário'])
            escritor.writeheader()
            escritor.writerows(horario)

    # Exportar horários por professor
    horarios_professores = defaultdict(list)

    for disciplina in nos:
        for prof in disciplina.professores:
            for dia in range(5):  # Semana tem 5 dias
                for turno in horarios[dia].keys():
                    if horarios[dia][turno] == disciplina.cor:
                        horarios_professores[prof].append({
                            'Curso': disciplina.curso,
                            'Período': disciplina.periodo,
                            'Código da Disciplina': disciplina.codigo,
                            'Nome da Disciplina': disciplina.nome,
                            'Dia da Semana': dia + 1,
                            'Professor': prof,
                            'Horário': turno
                        })

    # Escrever horários dos professores
    for prof, horario in horarios_professores.items():
        with open(os.path.join("professor/csv", f"horario_professor_{prof}.csv"), 'w', newline='', encoding='utf-8') as f:
            escritor = csv.DictWriter(f, fieldnames=['Curso', 'Período', 'Código da Disciplina',
                                                     'Nome da Disciplina', 'Dia da Semana',
                                                     'Professor', 'Horário'])
            escritor.writeheader()
            escritor.writerows(horario)

    # Horário global
    with open(os.path.join("controle", "horario_global.csv"), 'w', newline='', encoding='utf-8') as f:
        escritor = csv.DictWriter(f, fieldnames=['Curso', 'Período', 
                                                 'Código da Disciplina', 
                                                 'Nome da Disciplina', 
                                                 'Dia da Semana', 
                                                 'Professores', 
                                                 'Horário'])
        escritor.writeheader()
        for disciplina in nos:
            for dia in range(5):
                for turno in horarios[dia].keys():
                    if horarios[dia][turno] == disciplina.cor:
                        escritor.writerow({
                            'Curso': disciplina.curso,
                            'Período': disciplina.periodo,
                            'Código da Disciplina': disciplina.codigo,
                            'Nome da Disciplina': disciplina.nome,
                            'Dia da Semana': dia + 1,
                            'Professores': ', '.join(map(str, disciplina.professores)),
                            'Horário': turno
                        })

def processo_agendamento_principal():
    caminho_csv = os.path.join("..", "datasets", "csv", "semestre1.csv")
    nos = carregarDisciplinasCsv(caminho_csv)
    arestas = criarListaAdjacencia(nos)
    
    tentativas_maximas = 250

    for tentativa in range(tentativas_maximas):
        cores = colorirGrafoDSatur(nos, arestas)
        horarios = fazerDivisaoHorario(nos, arestas)

        if horarios is None:
            logging.error(f"Falha no agendamento na tentativa {tentativa}")
            continue

        if (verificar_restricoes_professor(horarios, nos) and
                verificar_restricoes_turma(horarios, nos) and
                verificar_restricoes_curso(horarios, nos)):
            
            # Criar e salvar o grafo de disciplinas e professores
            grafo_disciplina_curso = criar_grafo_disciplina_curso(nos)
            salvar_grafo_como_imagem(grafo_disciplina_curso)
            
            # Criar e salvar o grafo de restrições
            grafo_restricoes = criar_grafo_coloracao_restricoes(nos, arestas)
            salvar_grafo_restricoes_como_imagem(grafo_restricoes)

            # Exportar horários
            exportar_horarios(horarios, nos)

            return horarios, nos

    logging.critical("Não foi possível encontrar um agendamento válido após tentativas máximas")
    return None, None

def logicalXOR(a, b, condition):
    # return (a and not b) or (not a and b)
    return ((a == condition and b != condition) or (a != condition and b == condition))

def carregarDisciplinasCsv(caminhoCsv: str) -> list[Disciplina]:
    # Obtém o caminho absoluto do arquivo CSV
    caminho_absoluto = os.path.abspath(caminhoCsv)

    # Lê os dados do CSV
    with open(caminho_absoluto, encoding="utf-8") as arquivo:
        df = csv.reader(arquivo)
        # Lê o cabeçalho para obter a quantidade de colunas
        cabecalho = next(df)
        TOTAL_COLUNAS_CABECALHO = len(cabecalho)

        nos = []
        indice = 0
        for linha in df:
            professores = []
            for i in range(6, TOTAL_COLUNAS_CABECALHO, 1): # a partir do 6 tem o Prof 1
                if linha[i] == '1': # significa que o professor ministra a disciplina
                    professores.append(i - 5)
        
            curso = linha[0]
            ppc = linha[1]
            periodo = linha[2]
            codigo = linha[3]
            nome_disciplina = linha[4]
            ch = int(linha[5])

            
            if ch == 5:
                nos.append(Disciplina(indice, curso, ppc, periodo, codigo, nome_disciplina, 3, professores))
                nos.append(Disciplina(indice + 1, curso, ppc, periodo, codigo, nome_disciplina, 2, professores))
                indice += 2
            elif ch == 4:
                nos.append(Disciplina(indice, curso, ppc, periodo, codigo, nome_disciplina, 2, professores))
                nos.append(Disciplina(indice + 1, curso, ppc, periodo, codigo, nome_disciplina, 2, professores))
                indice += 2
            else:
                nos.append(Disciplina(indice, curso, ppc, periodo, codigo, nome_disciplina, ch, professores))
                indice += 1

        return nos


def criarListaAdjacencia(nos: list[Disciplina]) -> dict[int, set[int]]:
    listaAdjacencia = {}

    for i in range(len(nos)):

        arestas = set()

        for j in range(len(nos)):

            if i == j:
                continue

            # nao há porque permitir que nos com cargas horarias diferentes se conectem, ja que vao ser colocados em blocos diferentes de qualquer forma
            # if nos[i].ch != nos[j].ch:
            #     continue

            # se um dos cursos for sistemas e o outro nao for, nao precisa fazer as conexoes, ja que um é a noite e o outro nao
            if logicalXOR(nos[i].curso, nos[j].curso, 'SIN'):
                arestas.add(j)
            if nos[i].turma == nos[j].turma:
                arestas.add(j)

            for p in nos[i].professores:
                if p in nos[j].professores:
                    arestas.add(j)

        listaAdjacencia[i] = arestas

    return listaAdjacencia

def colorirGrafoDSatur(nos: list, listaAdjacencia: dict) -> int:
    # Inicialização
    cores_usadas = set()
    grau_saturacao = [0] * len(nos)
    cores_nos = [None] * len(nos)
    nao_coloridos = set(range(len(nos)))

    while nao_coloridos:
        # Seleciona o nó com maior grau de saturação
        # Em caso de empate, escolhe o de maior grau (número de vizinhos)
        max_saturacao = -1
        candidatos = []
        for i in nao_coloridos:
            sat = len(set(cores_nos[vizinho] for vizinho in listaAdjacencia[i] if cores_nos[vizinho] is not None))
            if sat > max_saturacao:
                max_saturacao = sat
                candidatos = [i]
            elif sat == max_saturacao:
                candidatos.append(i)
        # Seleciona o nó com maior grau entre os candidatos
        grau_max = -1
        for i in candidatos:
            grau = len(listaAdjacencia[i])
            if grau > grau_max:
                grau_max = grau
                no_escolhido = i
        # Atribui a menor cor possível
        cores_vizinhos = set(cores_nos[vizinho] for vizinho in listaAdjacencia[no_escolhido] if cores_nos[vizinho] is not None)
        cor = 0
        while cor in cores_vizinhos:
            cor += 1
        cores_nos[no_escolhido] = cor
        cores_usadas.add(cor)
        nao_coloridos.remove(no_escolhido)
        # Atualiza o grau de saturação dos vizinhos
        for vizinho in listaAdjacencia[no_escolhido]:
            if cores_nos[vizinho] is None:
                grau_saturacao[vizinho] += 1
    # Atribui as cores aos nós
    for i in range(len(nos)):
        nos[i].cor = cores_nos[i]
    return len(cores_usadas)

def colorirGrafo(nos: list, listaAdjacencia: dict) -> int:
    # Número de cores usadas até agora
    cores = 0

    # Para cada nó, vamos tentar colori-lo
    for i in range(len(nos)):

        # Lista de cores que não podem ser usadas
        coresBloqueadas = []

        # Para cada aresta, adiciona sua cor à lista de cores bloqueadas
        for aresta in listaAdjacencia[i]:
            
            if nos[aresta].cor != None:

                coresBloqueadas.append(nos[aresta].cor)

        # Inicializa corFinal como -1 para buscar uma cor disponível
        corFinal = -1

        # Procura uma cor que não esteja na lista de cores bloqueadas
        for cor in range(cores):

            if cor not in coresBloqueadas:

                corFinal = cor
                break

        # Se nenhuma cor disponível foi encontrada, cria uma nova cor
        if corFinal == -1:
            corFinal = cores
            cores += 1
        
        # Atribui a cor ao nó
        nos[i].cor = corFinal

    return cores


def criarGrafoTurmas(nos: list) -> dict:

    grafoTurmas = {}

    for i in range(len(nos)):
        arestas = set()

        for j in range(len(nos)):

            if i == j:
                continue
            
            if nos[i].turma == nos[j].turma:
                arestas.add(j)
        
        grafoTurmas [i] = arestas

    return grafoTurmas 


def fazerDivisaoHorario(nos: list[Disciplina], grafoColorido: dict):
    # Define uma semente fixa para os números aleatórios
    random.seed(42)

    # Primeiro passo: Criar os horarios, que são dicionários com os horarios de cada dia da semana
    horarios = [{}, {}, {}, {}, {}]
    for i in range(5):
        horarios[i] = {'M123': None, 'M45': None, 'T12': None, 'T345': None, 'N12': None, 'N345': None}

    
    # Segundo passo: Separar em listas de carga horaria 3 e 2
    ch3: list[Disciplina] = []
    ch2: list[Disciplina] = []
    for no in nos:
        if no.ch == 3:
            ch3.append(no)
        else:
            ch2.append(no)
    # Vamos usar o metodo shuffle para aleatoriezar um pouco a sequencia de horarios, evitando 5 horarios seguidos da mesma disciplina
    random.shuffle(ch3)
    random.shuffle(ch2)
    
    # Terceiro passo: Se separar uma cor para cada horario, podemos simplesmente conectar as disciplinas a suas respectivas cores
    for i in range(len(ch3)):

        # Uma condição de loop é estabelecida para captar erros onde um horario não foi encontrado
        sucesso = False
        while not sucesso:

            for dia in range(5):

                if ch3[i].curso == 'SIN':
                    if (horarios[dia]['N345'] == None or horarios[dia]['N345'] == ch3[i].cor):
                        horarios[dia]['N345'] = ch3[i].cor
                        ch3[i].horario = 'N345'
                        sucesso = True
                
                else:
                    if horarios[dia]['T345'] == None or horarios[dia]['T345'] == ch3[i].cor:
                        horarios[dia]['T345'] = ch3[i].cor
                        ch3[i].horario = 'T345'
                        sucesso = True

                    elif horarios[dia]['M123'] == None or horarios[dia]['M123'] == ch3[i].cor:
                        horarios[dia]['M123'] = ch3[i].cor
                        ch3[i].horario = 'M123'
                        sucesso = True
                
                if sucesso:
                    break

            # se nao tiver achado um horario bom, tentamos outras cores até achar uma que não tenha
            if not sucesso:
                ch3[i].cor += 1

                # se a cor for maior que 20 (só há 20 horarios no turno semanal mais longo), então um erro ocorreu na hora de criar as disciplinas
                if ch2[i].cor > 20:
                    print('Horario nao adicionado : ' + ch3[i].nome, ch3[i].curso, ch3[i].cor, ch3[i].turma, ch3[i].ch)
                    return None

    # Uma condição de loop é estabelecida para captar erros onde um horario não foi encontrado
    sucesso = False
    while not sucesso:

        # Agora faço a busca para as disciplinas de carga horaria 2
        for i in range(len(ch2)):
            sucesso = False

            # Primeiro procuro se a cor já está no horario
            for dia in range(5):

                if ch2[i].curso == 'SIN':
                    if (horarios[dia]['N345'] == ch2[i].cor):
                        ch2[i].horario = 'N34'
                        sucesso = True

                    elif (horarios[dia]['N12'] == ch2[i].cor):
                        ch2[i].horario = 'N12'
                        sucesso = True

                
                else:
                    if horarios[dia]['T345'] == ch2[i].cor:
                        ch2[i].horario = 'T34'
                        sucesso = True

                    elif horarios[dia]['T12'] == ch2[i].cor:
                        ch2[i].horario = 'T12'
                        sucesso = True


                    elif horarios[dia]['M123'] == ch2[i].cor:
                        ch2[i].horario = 'M12'
                        sucesso = True

                    elif horarios[dia]['M45'] == ch2[i].cor:
                        ch2[i].horario = 'M45'
                        sucesso = True

                # se ja tiver achado a cor, quebra o loop
                if sucesso:
                    break

            # se nao tiver achado o numero, coloca no primeiro None que achar
            if not sucesso:

                for dia in range(5):

                    if ch2[i].curso == 'SIN':
                        if (horarios[dia]['N345'] == None):
                            horarios[dia]['N345'] = ch2[i].cor
                            ch2[i].horario = 'N34'
                            sucesso = True

                        elif (horarios[dia]['N12'] == None):
                            horarios[dia]['N12'] = ch2[i].cor
                            ch2[i].horario = 'N12'
                            sucesso = True

                
                    else:
                        if horarios[dia]['T345'] == None:
                            horarios[dia]['T345'] = ch2[i].cor
                            ch2[i].horario = 'T34'
                            sucesso = True

                        elif horarios[dia]['T12'] == None:
                            horarios[dia]['T12'] = ch2[i].cor
                            ch2[i].horario = 'T12'
                            sucesso = True


                        elif horarios[dia]['M123'] == None:
                            horarios[dia]['M123'] = ch2[i].cor
                            ch2[i].horario = 'M12'
                            sucesso = True

                        elif horarios[dia]['M45'] == None:
                            horarios[dia]['M45'] = ch2[i].cor
                            ch2[i].horario = 'M45'
                            sucesso = True
                    
                    # se ja tiver achado a cor, quebra o loop
                    if sucesso:
                        break


            # se nao tiver achado um horario bom, tentamos outras cores até achar uma que não tenha
            if not sucesso:
                ch3[i].cor += 1

                # se a cor for maior que 20 (só há 20 horarios no turno semanal mais longo), então um erro ocorreu na hora de criar as disciplinas
                if ch2[i].cor > 20:
                    print('Horario nao adicionado : ' + ch2[i].nome, ch2[i].curso, ch2[i].cor, ch2[i].turma, ch2[i].ch)
                    return None
                
    return horarios

def exibirHorariosPorTurma(horarios, nos):
    dias = ["Segunda", "Terça", "Quarta", "Quinta", "Sexta"]
    turnos = ['M123', 'M45', 'T12', 'T345', 'N12', 'N345']

    # Agrupa as disciplinas por turma
    disciplinas_por_turma = defaultdict(list)
    for no in nos:
        disciplinas_por_turma[no.turma].append(no)

    # Itera por cada turma e exibe os horários
    for turma, disciplinas in disciplinas_por_turma.items():
        print(f"\nHorários para a turma {turma}:")
        
        for dia, horario in enumerate(horarios):
            print(f"\n  {dias[dia]}:")
            for turno in turnos:
                cor = horario[turno]
                if cor is not None:
                    # Filtra as disciplinas da turma que têm a mesma cor
                    disciplinas_no_turno = [
                        no.nome for no in disciplinas if no.cor == cor
                    ]
                    if disciplinas_no_turno:
                        print(f"    {turno}: {', '.join(disciplinas_no_turno)}")
                    else:
                        print(f"    {turno}: Livre")
                else:
                    print(f"    {turno}: Livre")

# Exibe os horários organizados por turma

def gerar_planilha_horarios(diretorio_csv, diretorio_xlsx, modo='professor'):
    """
    Gera planilhas de horários a partir de arquivos CSV.
    
    Parâmetros:
    - diretorio_csv: Diretório de origem dos arquivos CSV
    - diretorio_xlsx: Diretório de destino das planilhas Excel
    - modo: 'professor' ou 'aluno' (padrão: 'professor')
    """
    # Configurações de estilo
    preenchimento_cabecalho = PatternFill(start_color="4b598b", end_color="4b598b", fill_type="solid")
    fonte_cabecalho = Font(color="FFFFFF", bold=True)
    preenchimento_subcabecalho = PatternFill(start_color="333366", end_color="333366", fill_type="solid")
    fonte_subcabecalho = Font(color="FFFFFF", bold=True)
    preenchimento_linha = PatternFill(start_color="f9fbfd", end_color="f9fbfd", fill_type="solid")
    fonte_linha = Font(color="000000")
    alinhamento_centro = Alignment(horizontal="center", vertical="center")

    def processar_horario(horario, dia, horarios, codigo_disciplina):
        """Processa os horários das disciplinas"""
        mapeamento = {
            "M123": (0, 3),
            "M45": (3, 5),
            "T12": (5, 7),
            "T345": (7, 10),
            "N12": (10, 12),
            "N345": (12, 15),
        }
        inicio, fim = mapeamento[horario]
        for i in range(inicio, fim):
            horarios[i][dia] = codigo_disciplina

    # Cria o diretório de saída, se não existir
    os.makedirs(diretorio_xlsx, exist_ok=True)

    # Processa todos os arquivos CSV na pasta especificada
    for nome_arquivo in os.listdir(diretorio_csv):
        if nome_arquivo.endswith(".csv"):
            caminho_arquivo = os.path.join(diretorio_csv, nome_arquivo)
            
            # Lê o arquivo CSV
            with open(caminho_arquivo, mode="r", encoding="utf-8") as arquivo:
                leitor = csv.DictReader(arquivo)
                dados = list(leitor)

            # Cria um novo workbook
            pasta_trabalho = openpyxl.Workbook()

            if modo == 'aluno':
                # Itera por curso e período para alunos
                curso_periodo = {}
                for linha in dados:
                    chave = (linha["Curso"], linha["Período"])
                    if chave not in curso_periodo:
                        curso_periodo[chave] = []
                    curso_periodo[chave].append(linha)

                # Processa cada curso e período
                for (curso, periodo), linhas in curso_periodo.items():
                    planilha = pasta_trabalho.create_sheet(title=f"{curso} {periodo}")
                    pasta_trabalho.active = planilha

                    # Cabeçalho principal
                    planilha.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
                    planilha["A1"] = f"{curso} {periodo}"
                    planilha["A1"].fill = preenchimento_cabecalho
                    planilha["A1"].font = fonte_cabecalho
                    planilha["A1"].alignment = alinhamento_centro

                    # Subcabeçalho e configurações básicas
                    dias = ["Horários", "Seg", "Ter", "Qua", "Qui", "Sex"]
                    for num_coluna, dia in enumerate(dias, start=1):
                        celula = planilha.cell(row=2, column=num_coluna, value=dia)
                        celula.fill = preenchimento_subcabecalho
                        celula.font = fonte_subcabecalho
                        celula.alignment = alinhamento_centro

                    # Configurações de coluna
                    for coluna in range(1, 7):
                        planilha.column_dimensions[openpyxl.utils.get_column_letter(coluna)].width = 35

                    # Adiciona horários
                    horarios = [
                        "07:00 - 07:55", "07:55 - 08:50", "08:50 - 09:45",
                        "10:10 - 11:05", "11:05 - 12:00",
                        "13:30 - 14:25", "14:25 - 15:20", "15:45 - 16:40", "16:40 - 17:35", "17:35 - 18:30",
                        "19:00 - 19:50", "19:50 - 20:40", "21:00 - 21:50", "21:50 - 22:40", "22:40 - 23:30"
                    ]
                    linhas_horario = [["-- -- -- --"] * 5 for _ in horarios]

                    for num_linha, horario in enumerate(horarios, start=3):
                        planilha.cell(row=num_linha, column=1, value=horario).fill = preenchimento_linha
                        planilha.cell(row=num_linha, column=1).font = fonte_linha
                        planilha.cell(row=num_linha, column=1).alignment = alinhamento_centro

                    # Popula horários e legenda
                    legenda = {}
                    for linha in linhas:
                        dia = int(linha["Dia da Semana"])
                        if dia < 1 or dia > 5:
                            continue
                        dia -= 1

                        codigo = linha["Código da Disciplina"]
                        processar_horario(linha["Horário"], dia, linhas_horario, codigo)

                        if codigo not in legenda:
                            legenda[codigo] = f"{codigo}: {linha['Nome da Disciplina']} - {linha['Professor']}"

                    # Preenche planilha
                    for num_linha, horario in enumerate(linhas_horario, start=3):
                        for num_coluna, valor in enumerate(horario, start=2):
                            celula = planilha.cell(row=num_linha, column=num_coluna, value=valor)
                            celula.fill = preenchimento_linha
                            celula.font = fonte_linha
                            celula.alignment = alinhamento_centro

                    # Adiciona legenda
                    linha_legenda_inicio = len(horarios) + 4
                    planilha.cell(row=linha_legenda_inicio, column=1, value="Legenda:").font = Font(bold=True)
                    for i, texto in enumerate(legenda.values(), start=linha_legenda_inicio + 1):
                        planilha.cell(row=i, column=1, value=texto)

                # Remove aba padrão
                pasta_trabalho.remove(pasta_trabalho["Sheet"])

            else:  # modo professor
                planilha = pasta_trabalho.active
                planilha.title = "Horários"

                # Cabeçalho principal
                planilha.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
                planilha["A1"] = "Horários de Aulas"
                planilha["A1"].fill = preenchimento_cabecalho
                planilha["A1"].font = fonte_cabecalho
                planilha["A1"].alignment = alinhamento_centro

                # Subcabeçalho
                dias = ["Horários", "Seg", "Ter", "Qua", "Qui", "Sex"]
                for num_coluna, dia in enumerate(dias, start=1):
                    celula = planilha.cell(row=2, column=num_coluna, value=dia)
                    celula.fill = preenchimento_subcabecalho
                    celula.font = fonte_subcabecalho
                    celula.alignment = alinhamento_centro

                # Configurações de coluna
                for coluna in range(1, 7):
                    planilha.column_dimensions[openpyxl.utils.get_column_letter(coluna)].width = 35

                # Adiciona horários
                horarios = [
                    "07:00 - 07:55", "07:55 - 08:50", "08:50 - 09:45",
                    "10:10 - 11:05", "11:05 - 12:00",
                    "13:30 - 14:25", "14:25 - 15:20", "15:45 - 16:40", "16:40 - 17:35", "17:35 - 18:30",
                    "19:00 - 19:50", "19:50 - 20:40", "21:00 - 21:50", "21:50 - 22:40", "22:40 - 23:30"
                ]
                linhas_horario = [["-- -- -- --"] * 5 for _ in horarios]

                for num_linha, horario in enumerate(horarios, start=3):
                    planilha.cell(row=num_linha, column=1, value=horario).fill = preenchimento_linha
                    planilha.cell(row=num_linha, column=1).font = fonte_linha
                    planilha.cell(row=num_linha, column=1).alignment = alinhamento_centro

                # Popula horários e legenda
                legenda = {}
                for linha in dados:
                    dia = int(linha["Dia da Semana"])
                    if dia < 1 or dia > 5:
                        continue
                    dia -= 1

                    codigo = f"{linha['Código da Disciplina']} - {linha['Curso']} - {linha['Período']}"
                    processar_horario(linha["Horário"], dia, linhas_horario, codigo)

                    if codigo not in legenda:
                        legenda[codigo] = f"{codigo}: {linha['Nome da Disciplina']} - {linha['Curso']} - {linha['Período']}"

                # Preenche planilha
                for num_linha, horario in enumerate(linhas_horario, start=3):
                    for num_coluna, valor in enumerate(horario, start=2):
                        celula = planilha.cell(row=num_linha, column=num_coluna, value=valor)
                        celula.fill = preenchimento_linha
                        celula.font = fonte_linha
                        celula.alignment = alinhamento_centro

                # Adiciona legenda
                linha_legenda_inicio = len(horarios) + 4
                planilha.cell(row=linha_legenda_inicio, column=1, value="Legenda:").font = Font(bold=True)
                for i, texto in enumerate(legenda.values(), start=linha_legenda_inicio + 1):
                    planilha.cell(row=i, column=1, value=texto)

            # Salva o arquivo Excel
            caminho_saida = os.path.join(diretorio_xlsx, f"{os.path.splitext(nome_arquivo)[0]}.xlsx")
            pasta_trabalho.save(caminho_saida)
            print(f"Arquivo {caminho_saida} salvo com sucesso!")

if __name__ == "__main__":
    # Constrói o caminho para o arquivo CSV de forma compatível com qualquer sistema operacional
    #caminho_csv = os.path.join("..", "datasets", "csv", "semestre1.csv")

    # Chama a função com o caminho ajustado
    #nos = carregarDisciplinasCsv(caminho_csv)
    # arestas = fazerArestas(nos)
    #arestas = criarListaAdjacencia(nos)

    # print(arestas)
    #    print(f'Quantidade de cores: {colorirGrafo(nos, arestas)}')

    #    grafoPorTurmas = criarGrafoTurmas(nos)
    # print(grafoPorTurmas)

    #    horarios = fazerDivisaoHorario(nos, arestas)

    # print(horarios)

    # Exibe os horários formatados
    #    exibirHorariosPorTurma(horarios, nos)

    processo_agendamento_principal()
    gerar_planilha_horarios('professor/csv', 'professor/xlsx', modo='professor')
    gerar_planilha_horarios('aluno/csv', 'aluno/xlsx', modo='aluno')